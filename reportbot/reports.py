"""Report rendering: text summary, CSV export and an XLSX workbook."""

from __future__ import annotations

import csv
import io
from datetime import datetime

from .parsing import format_money
from .store import Expense


def render_summary(totals: list[tuple[str, int]], start: str, end: str,
                   currency: str = "EUR", width: int = 12) -> str:
    """A monospace breakdown with a simple bar chart, readable on a phone."""
    if not totals:
        return f"No expenses between {start} and {end}."

    grand_total = sum(amount for _, amount in totals)
    longest = max(len(category) for category, _ in totals)
    lines = [f"Expenses {start} to {end}", ""]

    for category, amount in totals:
        share = amount / grand_total
        bar = "█" * max(1, round(share * width))
        lines.append(
            f"{category.ljust(longest)}  {format_money(amount, currency).rjust(10)}  "
            f"{bar} {share * 100:4.1f}%"
        )

    lines += ["", f"{'total'.ljust(longest)}  {format_money(grand_total, currency).rjust(10)}"]
    return "\n".join(lines)


def render_csv(expenses: list[Expense]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["date", "category", "amount", "currency", "note"])
    for item in expenses:
        writer.writerow([item.spent_on, item.category, f"{item.amount:.2f}",
                         item.currency, item.note])
    return buffer.getvalue().encode("utf-8-sig")  # BOM so Excel opens it correctly


def render_xlsx(expenses: list[Expense], totals: list[tuple[str, int]],
                start: str, end: str) -> bytes:
    """Build a two-sheet workbook: line items plus a summary with a chart."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import PieChart, Reference
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "XLSX export needs openpyxl. Install it with: pip install '.[xlsx]'"
        ) from exc

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Expenses"
    headers = ["Date", "Category", "Amount", "Currency", "Note"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for item in expenses:
        sheet.append([item.spent_on, item.category, item.amount, item.currency, item.note])

    for column, header in enumerate(headers, start=1):
        longest = max([len(header)] + [
            len(str(sheet.cell(row=r, column=column).value or ""))
            for r in range(2, sheet.max_row + 1)
        ])
        sheet.column_dimensions[get_column_letter(column)].width = min(longest + 2, 50)
    sheet.freeze_panes = "A2"

    summary = workbook.create_sheet("Summary")
    summary["A1"] = f"Expenses {start} to {end}"
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(["Category", "Amount"])
    for cell in summary[3]:
        cell.font = Font(bold=True)

    for category, amount in totals:
        summary.append([category, amount / 100])

    total_row = summary.max_row + 1
    summary.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    summary.cell(row=total_row, column=2,
                 value=f"=SUM(B4:B{total_row - 1})").font = Font(bold=True)

    if totals:
        chart = PieChart()
        chart.title = "Share by category"
        chart.add_data(Reference(summary, min_col=2, min_row=3, max_row=total_row - 1),
                       titles_from_data=True)
        chart.set_categories(Reference(summary, min_col=1, min_row=4, max_row=total_row - 1))
        summary.add_chart(chart, "D3")

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 14

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def report_filename(prefix: str, start: str, end: str, extension: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d")
    return f"{prefix}_{start}_{end}_{stamp}.{extension}"
